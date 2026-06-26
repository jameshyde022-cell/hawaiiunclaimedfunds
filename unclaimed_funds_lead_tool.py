#!/usr/bin/env python3
"""
Build a Hawaii unclaimed-funds lead workbook from scraped or pasted records.

This tool does not scrape websites, bypass CAPTCHA, or identify a final contact
person. It cleans records, groups likely related owners, and creates research
queues that require human verification.
"""

from __future__ import annotations

import argparse
import html
import io
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus

import pandas as pd

try:
    from rapidfuzz import fuzz
except ImportError:  # pragma: no cover - useful when reading without deps.
    from difflib import SequenceMatcher

    class fuzz:  # type: ignore[no-redef]
        @staticmethod
        def token_sort_ratio(left: str, right: str) -> float:
            left_tokens = " ".join(sorted(left.split()))
            right_tokens = " ".join(sorted(right.split()))
            return SequenceMatcher(None, left_tokens, right_tokens).ratio() * 100


STANDARD_COLUMNS = {
    "reported owner": "Reported Owner",
    "owner": "Reported Owner",
    "name": "Reported Owner",
    "owner name": "Reported Owner",
    "co-owner": "Co-owner",
    "co owner": "Co-owner",
    "coowner": "Co-owner",
    "joint owner": "Co-owner",
    "address": "Address / Location",
    "location": "Address / Location",
    "address / location": "Address / Location",
    "last known address": "Address / Location",
    "reporting company": "Reporting Company",
    "holder": "Reporting Company",
    "holder name": "Reporting Company",
    "company": "Reporting Company",
    "cash amount": "Cash Amount",
    "amount": "Cash Amount",
    "value": "Cash Amount",
    "cash": "Cash Amount",
    "shares": "Shares",
    "stock shares": "Shares",
    "property id": "Property ID",
    "property id if available": "Property ID",
    "id": "Property ID",
    "source": "Source",
    "source if available": "Source",
}

OUTPUT_COLUMNS = [
    "Record ID",
    "Group ID",
    "Group Confidence",
    "Reported Owner",
    "Normalized Owner",
    "Co-owner",
    "Address / Location",
    "Reporting Company",
    "Cash Amount",
    "Shares",
    "Property ID",
    "Source",
    "Record Count In Group",
    "Grouped Total Cash",
    "Estimated Recovery Fee 20%",
    "Priority",
    "Lead Type Tags",
    "Research Needed",
    "Needs Human Verification",
    "Owner Search Link",
    "DCCA Search Link",
]

BUSINESS_TERMS = {
    "LLC",
    "L L C",
    "INC",
    "CORP",
    "CORPORATION",
    "CO",
    "COMPANY",
    "LTD",
    "LIMITED",
    "LP",
    "LLP",
    "PLLC",
    "DBA",
    "ASSOC",
    "ASSOCIATION",
    "FOUNDATION",
    "PARTNERS",
    "PARTNERSHIP",
    "ENTERPRISES",
    "GROUP",
    "SERVICES",
    "HOLDINGS",
    "RESTAURANT",
    "AUTO",
    "REPAIR",
    "CONSTRUCTION",
    "REALTY",
    "PROPERTIES",
}

GOVERNMENT_TERMS = {
    "STATE OF",
    "CITY AND COUNTY",
    "COUNTY OF",
    "DEPARTMENT",
    "DIVISION",
    "BOARD OF",
    "AUTHORITY",
    "UNIVERSITY OF HAWAII",
    "PUBLIC",
    "SCHOOL",
}

ESCROW_TRUST_TERMS = {"ESCROW", "TRUST", "TRUSTEE", "ESTATE", "FIDUCIARY", "CUSTODIAN"}
INSURANCE_TERMS = {"INSURANCE", "ASSURANCE", "INDEMNITY", "MUTUAL", "CASUALTY", "LIFE INS"}
BANK_TERMS = {"BANK", "CREDIT UNION", "SAVINGS", "CD", "CERTIFICATE OF DEPOSIT", "FINANCIAL"}
COURT_TERMS = {"COURT", "JUDICIAL", "CLERK", "CASE", "SETTLEMENT", "GARNISH", "PROBATE"}

NAME_NOISE = {
    "MR",
    "MRS",
    "MS",
    "MISS",
    "DR",
    "PROF",
    "REV",
    "THE",
    "ATTN",
    "C O",
    "CARE OF",
}


@dataclass
class GroupStats:
    group_id: str
    total_cash: float
    record_count: int
    confidence: str


@dataclass
class ReportSection:
    title: str
    filename: str
    dataframe: pd.DataFrame


class UnionFind:
    def __init__(self, size: int) -> None:
        self.parent = list(range(size))

    def find(self, item: int) -> int:
        while self.parent[item] != item:
            self.parent[item] = self.parent[self.parent[item]]
            item = self.parent[item]
        return item

    def union(self, left: int, right: int) -> None:
        left_root = self.find(left)
        right_root = self.find(right)
        if left_root != right_root:
            self.parent[right_root] = left_root


def clean_header(header: object) -> str:
    return re.sub(r"\s+", " ", str(header).strip().lower())


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    renamed = {}
    for column in df.columns:
        key = clean_header(column)
        renamed[column] = STANDARD_COLUMNS.get(key, str(column).strip())
    df = df.rename(columns=renamed)

    for column in [
        "Reported Owner",
        "Co-owner",
        "Address / Location",
        "Reporting Company",
        "Cash Amount",
        "Shares",
        "Property ID",
        "Source",
    ]:
        if column not in df.columns:
            df[column] = ""

    return df


def parse_money(value: object) -> float:
    if pd.isna(value):
        return 0.0
    text = str(value).strip()
    if not text:
        return 0.0
    if text.lower() in {"unknown", "undisclosed", "n/a", "na"}:
        return 0.0
    negative = text.startswith("(") and text.endswith(")")
    cleaned = re.sub(r"[^0-9.\-]", "", text)
    if cleaned in {"", ".", "-", "-."}:
        return 0.0
    try:
        amount = float(cleaned)
    except ValueError:
        return 0.0
    return -amount if negative else amount


def normalize_owner_name(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value).upper()
    text = text.replace("&", " AND ")
    text = re.sub(r"[^A-Z0-9\s]", " ", text)
    text = re.sub(r"\b(JR|SR|II|III|IV|V)\b", " ", text)
    for phrase in sorted(NAME_NOISE, key=len, reverse=True):
        text = re.sub(rf"\b{re.escape(phrase)}\b", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def grouping_key(name: str) -> str:
    tokens = name.split()
    if not tokens:
        return ""
    non_business = [token for token in tokens if token not in BUSINESS_TERMS]
    if len(non_business) >= 2:
        return non_business[0][:4] + "|" + non_business[-1][:4]
    return tokens[0][:6]


def fuzzy_group_records(df: pd.DataFrame, threshold: int) -> tuple[pd.Series, pd.Series]:
    names = df["Normalized Owner"].fillna("").astype(str).tolist()
    keys = [grouping_key(name) for name in names]
    uf = UnionFind(len(names))
    matched_scores: dict[int, list[float]] = {index: [] for index in range(len(names))}

    buckets: dict[str, list[int]] = {}
    for index, key in enumerate(keys):
        buckets.setdefault(key, []).append(index)

    for indexes in buckets.values():
        for offset, left in enumerate(indexes):
            for right in indexes[offset + 1 :]:
                if not names[left] or not names[right]:
                    continue
                score = float(fuzz.token_sort_ratio(names[left], names[right]))
                if score >= threshold:
                    uf.union(left, right)
                    matched_scores[left].append(score)
                    matched_scores[right].append(score)

    roots = [uf.find(index) for index in range(len(names))]
    root_to_group: dict[int, str] = {}
    group_ids = []
    for root in roots:
        if root not in root_to_group:
            root_to_group[root] = f"G{len(root_to_group) + 1:04d}"
        group_ids.append(root_to_group[root])

    confidence = []
    for index, group_id in enumerate(group_ids):
        group_size = group_ids.count(group_id)
        if group_size == 1:
            confidence.append("single record")
        elif matched_scores[index]:
            confidence.append(f"fuzzy {round(max(matched_scores[index]))}")
        else:
            confidence.append("grouped")

    return pd.Series(group_ids, index=df.index), pd.Series(confidence, index=df.index)


def contains_any(text: str, terms: Iterable[str]) -> bool:
    return any(term in text for term in terms)


def lead_tags(row: pd.Series) -> list[str]:
    owner = str(row.get("Normalized Owner", ""))
    co_owner = str(row.get("Co-owner", "")).strip()
    holder = normalize_owner_name(row.get("Reporting Company", ""))
    combined = f"{owner} {holder}"

    tags: list[str] = []
    if contains_any(owner, BUSINESS_TERMS):
        tags.append("business")
    if co_owner:
        tags.append("co-owner")
    if contains_any(combined, GOVERNMENT_TERMS):
        tags.append("government")
    if contains_any(combined, ESCROW_TRUST_TERMS):
        tags.append("escrow/trust")
    if contains_any(combined, INSURANCE_TERMS):
        tags.append("insurance-related")
    if contains_any(combined, BANK_TERMS):
        tags.append("bank/CD")
    if contains_any(combined, COURT_TERMS):
        tags.append("court-related")
    if "business" not in tags and "government" not in tags:
        tags.insert(0, "individual")
    return tags


def research_steps(tags: str, owner: str, amount: float, grouped_total: float) -> str:
    tag_set = {tag.strip() for tag in tags.split(";") if tag.strip()}
    steps: list[str] = []

    if "business" in tag_set:
        steps.extend(
            [
                "Check Hawaii DCCA",
                "Find registered agent",
                "Verify business status",
                "Find owner/manager",
                "Search company website",
            ]
        )
    if "individual" in tag_set:
        steps.extend(["Verify address", "Search LinkedIn"])
    if "co-owner" in tag_set:
        steps.append("Confirm all co-owner documentation requirements")
    if "government" in tag_set:
        steps.append("Call State Unclaimed Property for document requirements")
    if {"escrow/trust", "court-related"} & tag_set:
        steps.append("Review trust, estate, court, or settlement document requirements")
    if {"insurance-related", "bank/CD"} & tag_set:
        steps.append("Confirm holder-specific proof requirements")
    if amount >= 1000 or grouped_total >= 1000:
        steps.append("Prioritize outreach after human verification")
    if not owner:
        steps.append("Review source record because owner name is blank")

    deduped = list(dict.fromkeys(steps))
    return "; ".join(deduped)


def priority(total_cash: float, record_count: int) -> str:
    if total_cash >= 5000:
        return "High"
    if total_cash >= 1000:
        return "Medium"
    if record_count > 1:
        return "Low - grouped"
    return "Low"


def search_link(query: str) -> str:
    if not query:
        return ""
    return f"https://www.google.com/search?q={quote_plus(query)}"


def dcca_link(owner: str) -> str:
    if not owner:
        return ""
    return f"https://hbe.ehawaii.gov/documents/search.html?search={quote_plus(owner)}"


def read_input(path: Path, text_mode: bool = False) -> pd.DataFrame:
    if text_mode:
        text = path.read_text(encoding="utf-8-sig")
        return read_pasted_text(text)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path)
    if path.suffix.lower() in {".csv", ".txt", ".tsv"}:
        if path.suffix.lower() == ".tsv":
            return pd.read_csv(path, sep="\t")
        return pd.read_csv(path)
    raise ValueError("Input must be .csv, .xlsx, .xls, .xlsm, .txt, or .tsv")


def read_pasted_text(text: str) -> pd.DataFrame:
    stripped = text.strip()
    if not stripped:
        raise ValueError("Pasted text input is empty")
    first_line = stripped.splitlines()[0]
    if "\t" in first_line:
        return pd.read_csv(io.StringIO(stripped), sep="\t")
    return pd.read_csv(io.StringIO(stripped), sep=None, engine="python")


def build_lead_dataframe(raw_df: pd.DataFrame, fuzzy_threshold: int) -> pd.DataFrame:
    df = normalize_columns(raw_df.copy())
    df = df.reset_index(drop=True)
    df["Record ID"] = [f"R{index + 1:05d}" for index in range(len(df))]

    for column in ["Reported Owner", "Co-owner", "Address / Location", "Reporting Company", "Property ID", "Source"]:
        df[column] = df[column].fillna("").astype(str).str.strip()

    df["Cash Amount"] = df["Cash Amount"].apply(parse_money)
    df["Shares"] = df["Shares"].fillna("").astype(str).str.strip()
    df["Normalized Owner"] = df["Reported Owner"].apply(normalize_owner_name)

    group_ids, group_confidence = fuzzy_group_records(df, fuzzy_threshold)
    df["Group ID"] = group_ids
    df["Group Confidence"] = group_confidence

    group_stats: dict[str, GroupStats] = {}
    grouped = df.groupby("Group ID", dropna=False)
    for group_id, group_df in grouped:
        total = float(group_df["Cash Amount"].sum())
        count = int(len(group_df))
        if count == 1:
            confidence = "single record"
        else:
            confidence = ", ".join(sorted(set(group_df["Group Confidence"])))
        group_stats[str(group_id)] = GroupStats(str(group_id), total, count, confidence)

    df["Grouped Total Cash"] = df["Group ID"].map(lambda gid: group_stats[str(gid)].total_cash)
    df["Record Count In Group"] = df["Group ID"].map(lambda gid: group_stats[str(gid)].record_count)
    df["Estimated Recovery Fee 20%"] = df["Grouped Total Cash"] * 0.20
    df["Priority"] = df.apply(lambda row: priority(float(row["Grouped Total Cash"]), int(row["Record Count In Group"])), axis=1)
    df["Lead Type Tags"] = df.apply(lambda row: ";".join(lead_tags(row)), axis=1)
    df["Research Needed"] = df.apply(
        lambda row: research_steps(
            str(row["Lead Type Tags"]),
            str(row["Normalized Owner"]),
            float(row["Cash Amount"]),
            float(row["Grouped Total Cash"]),
        ),
        axis=1,
    )
    df["Needs Human Verification"] = "Yes"
    df["Owner Search Link"] = df.apply(
        lambda row: search_link(
            " ".join(
                part
                for part in [
                    str(row["Reported Owner"]),
                    str(row["Address / Location"]),
                    "Hawaii",
                ]
                if part and part.lower() != "nan"
            )
        ),
        axis=1,
    )
    df["DCCA Search Link"] = df["Reported Owner"].apply(dcca_link)

    extras = [column for column in df.columns if column not in OUTPUT_COLUMNS]
    return df[OUTPUT_COLUMNS + extras]


def grouped_owner_summary(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for group_id, group in df.groupby("Group ID", dropna=False):
        first = group.iloc[0]
        rows.append(
            {
                "Group ID": group_id,
                "Primary Owner": first["Reported Owner"],
                "Normalized Owner": first["Normalized Owner"],
                "Record Count": len(group),
                "Grouped Total Cash": float(group["Cash Amount"].sum()),
                "Max Single Cash": float(group["Cash Amount"].max()),
                "Estimated Recovery Fee 20%": float(group["Cash Amount"].sum()) * 0.20,
                "Priority": priority(float(group["Cash Amount"].sum()), len(group)),
                "Lead Type Tags": ";".join(sorted(set(";".join(group["Lead Type Tags"]).split(";")))),
                "Research Needed": "; ".join(dict.fromkeys("; ".join(group["Research Needed"]).split("; "))),
                "Needs Human Verification": "Yes",
                "Owner Search Link": first["Owner Search Link"],
                "DCCA Search Link": first["DCCA Search Link"],
                "Group Confidence": "; ".join(sorted(set(group["Group Confidence"]))),
                "Property IDs": "; ".join(str(value) for value in group["Property ID"].dropna().unique() if str(value).strip()),
                "Sources": "; ".join(str(value) for value in group["Source"].dropna().unique() if str(value).strip()),
            }
        )
    return pd.DataFrame(rows).sort_values(["Grouped Total Cash", "Record Count"], ascending=[False, False])


def dashboard(df: pd.DataFrame, grouped_df: pd.DataFrame) -> pd.DataFrame:
    high = grouped_df[grouped_df["Priority"] == "High"]
    medium = grouped_df[grouped_df["Priority"] == "Medium"]
    grouped_1000 = grouped_df[
        (grouped_df["Grouped Total Cash"] >= 1000)
        & (grouped_df["Record Count"] > 1)
        & (grouped_df["Max Single Cash"] < 1000)
    ]
    single_1000 = df[df["Cash Amount"] >= 1000]

    tag_counts: dict[str, int] = {}
    for tag_string in df["Lead Type Tags"]:
        for tag in str(tag_string).split(";"):
            if tag:
                tag_counts[tag] = tag_counts.get(tag, 0) + 1

    rows = [
        {"Metric": "Total cleaned records", "Value": len(df)},
        {"Metric": "Likely owner groups", "Value": len(grouped_df)},
        {"Metric": "Single claims >= $1,000", "Value": len(single_1000)},
        {"Metric": "Grouped owners >= $1,000 from multiple records", "Value": len(grouped_1000)},
        {"Metric": "High priority owner groups", "Value": len(high)},
        {"Metric": "Medium priority owner groups", "Value": len(medium)},
        {"Metric": "Total cash amount in input", "Value": round(float(df["Cash Amount"].sum()), 2)},
        {"Metric": "Estimated 20% fee on all grouped totals", "Value": round(float(grouped_df["Estimated Recovery Fee 20%"].sum()), 2)},
    ]
    rows.extend({"Metric": f"Records tagged {tag}", "Value": count} for tag, count in sorted(tag_counts.items()))
    rows.append({"Metric": "Compliance note", "Value": "Research links require human verification; tool does not scrape or bypass CAPTCHA."})
    return pd.DataFrame(rows)


def report_sections(df: pd.DataFrame) -> list[ReportSection]:
    grouped_df = grouped_owner_summary(df)
    single_1000 = df[df["Cash Amount"] >= 1000].sort_values("Cash Amount", ascending=False)
    grouped_1000 = grouped_df[
        (grouped_df["Grouped Total Cash"] >= 1000)
        & (grouped_df["Record Count"] > 1)
        & (grouped_df["Max Single Cash"] < 1000)
    ]
    high_priority = grouped_df[grouped_df["Priority"] == "High"]
    business = df[
        df["Lead Type Tags"].str.contains("business", na=False)
        & ~df["Lead Type Tags"].str.contains("co-owner", na=False)
    ]
    complex_claims = df[df["Lead Type Tags"].str.contains("co-owner|escrow/trust|court-related", regex=True, na=False)]
    research_queue = df[df["Research Needed"].astype(str).str.len() > 0].copy()
    priority_rank = {"High": 0, "Medium": 1, "Low - grouped": 2, "Low": 3}
    research_queue["_Priority Rank"] = research_queue["Priority"].map(priority_rank).fillna(9)
    research_queue = (
        research_queue.sort_values(["_Priority Rank", "Grouped Total Cash"], ascending=[True, False])
        .drop(columns=["_Priority Rank"])
    )

    return [
        ReportSection("Summary Dashboard", "summary_dashboard.csv", dashboard(df, grouped_df)),
        ReportSection("High Priority Leads", "high_priority_leads.csv", high_priority),
        ReportSection("Business Leads", "business_leads.csv", business),
        ReportSection("$1,000+ Single Claims", "single_claims_1000_plus.csv", single_1000),
        ReportSection("$1,000+ Grouped Owners", "grouped_owners_1000_plus.csv", grouped_1000),
        ReportSection("Co-owner Complex Claims", "co_owner_complex_claims.csv", complex_claims),
        ReportSection("Research Queue", "research_queue.csv", research_queue),
        ReportSection("All Cleaned Records", "all_cleaned_records.csv", df),
    ]


def autosize_workbook(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, min(len(value), 60))
            sheet.column_dimensions[column_letter].width = max(max_length + 2, 12)
        sheet.auto_filter.ref = sheet.dimensions
    workbook.save(path)


def write_workbook(df: pd.DataFrame, output_path: Path) -> None:
    sections_by_title = {section.title: section for section in report_sections(df)}
    workbook_order = [
        "All Cleaned Records",
        "$1,000+ Single Claims",
        "$1,000+ Grouped Owners",
        "High Priority Leads",
        "Business Leads",
        "Co-owner Complex Claims",
        "Research Queue",
        "Summary Dashboard",
    ]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for title in workbook_order:
            section = sections_by_title[title]
            section.dataframe.to_excel(writer, sheet_name=section.title, index=False)

    autosize_workbook(output_path)


def format_cell(value: object, column: str) -> str:
    if pd.isna(value):
        return ""
    if column in {"Cash Amount", "Grouped Total Cash", "Max Single Cash", "Estimated Recovery Fee 20%", "Value"}:
        try:
            amount = float(value)
        except (TypeError, ValueError):
            return html.escape(str(value))
        if column == "Value":
            return f"{amount:,.2f}" if not amount.is_integer() else f"{amount:,.0f}"
        return f"${amount:,.2f}"
    return html.escape(str(value))


def html_columns(df: pd.DataFrame, title: str) -> list[str]:
    if title == "Summary Dashboard":
        return [column for column in ["Metric", "Value"] if column in df.columns]
    preferred = [
        "Priority",
        "Reported Owner",
        "Primary Owner",
        "Co-owner",
        "Address / Location",
        "Cash Amount",
        "Grouped Total Cash",
        "Estimated Recovery Fee 20%",
        "Record Count",
        "Record Count In Group",
        "Max Single Cash",
        "Reporting Company",
        "Lead Type Tags",
        "Research Needed",
        "Needs Human Verification",
        "Group Confidence",
        "Property ID",
        "Property IDs",
    ]
    selected = [column for column in preferred if column in df.columns]
    return selected or list(df.columns)


def render_html_table(df: pd.DataFrame, title: str) -> str:
    columns = html_columns(df, title)
    header = "".join(f"<th>{html.escape(column)}</th>" for column in columns)
    if df.empty:
        body = f'<tr><td colspan="{len(columns)}">No records in this section.</td></tr>'
    else:
        rows = []
        for _, row in df.iterrows():
            cells = "".join(f"<td>{format_cell(row[column], column)}</td>" for column in columns)
            rows.append(f"<tr>{cells}</tr>")
        body = "\n".join(rows)
    return f"<table><thead><tr>{header}</tr></thead><tbody>{body}</tbody></table>"


def write_csv_outputs(sections: list[ReportSection], output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    for section in sections:
        section.dataframe.to_csv(output_dir / section.filename, index=False)


def write_html_report(sections: list[ReportSection], output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    summary = sections[0].dataframe if sections else pd.DataFrame()
    summary_rows = ""
    for _, row in summary.iterrows():
        metric = html.escape(str(row.get("Metric", "")))
        value = format_cell(row.get("Value", ""), "Value")
        summary_rows += f'<div class="total"><span>{metric}</span><strong>{value}</strong></div>\n'

    nav = "\n".join(
        f'<a href="#{html.escape(section.filename[:-4])}">{html.escape(section.title)}</a>'
        for section in sections
    )
    sections_html = "\n".join(
        f'''
        <section id="{html.escape(section.filename[:-4])}">
            <h2>{html.escape(section.title)}</h2>
            <p class="meta">{len(section.dataframe)} records. CSV: {html.escape(section.filename)}</p>
            {render_html_table(section.dataframe, section.title)}
        </section>
        '''
        for section in sections
    )
    report = f'''<!doctype html>
<html lang="en">
<head>
    <meta charset="utf-8">
    <meta name="viewport" content="width=device-width, initial-scale=1">
    <title>Unclaimed Funds Lead Report</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 24px; color: #1f2933; background: #f7f8fa; }}
        h1 {{ margin-bottom: 6px; }}
        h2 {{ margin-top: 34px; border-bottom: 2px solid #d8dee8; padding-bottom: 6px; }}
        .subtitle, .meta {{ color: #52606d; }}
        .totals {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(230px, 1fr)); gap: 10px; margin: 18px 0; }}
        .total {{ background: #ffffff; border: 1px solid #d8dee8; padding: 12px; }}
        .total span {{ display: block; color: #52606d; font-size: 13px; }}
        .total strong {{ display: block; margin-top: 5px; font-size: 18px; }}
        nav {{ display: flex; flex-wrap: wrap; gap: 8px; margin: 18px 0 24px; }}
        nav a {{ background: #ffffff; border: 1px solid #cbd5e1; color: #102a43; padding: 7px 10px; text-decoration: none; }}
        table {{ border-collapse: collapse; width: 100%; background: #ffffff; margin-bottom: 24px; }}
        th, td {{ border: 1px solid #d8dee8; padding: 7px 8px; text-align: left; vertical-align: top; font-size: 13px; }}
        th {{ background: #102a43; color: #ffffff; position: sticky; top: 0; }}
        tr:nth-child(even) td {{ background: #f3f6f9; }}
        section {{ overflow-x: auto; }}
    </style>
</head>
<body>
    <h1>Unclaimed Funds Lead Report</h1>
    <p class="subtitle">Generated from cleaned records. Every lead still requires human verification.</p>
    <div class="totals">
        {summary_rows}
    </div>
    <nav>
        {nav}
    </nav>
    {sections_html}
</body>
</html>
'''
    report_path = output_dir / "lead_report.html"
    report_path.write_text(report, encoding="utf-8")
    return report_path


def write_browser_outputs(df: pd.DataFrame, output_dir: Path) -> tuple[Path, list[ReportSection]]:
    sections = report_sections(df)
    write_csv_outputs(sections, output_dir)
    report_path = write_html_report(sections, output_dir)
    return report_path, sections


def main() -> int:
    parser = argparse.ArgumentParser(description="Create Hawaii unclaimed-funds lead reports from CSV, Excel, or pasted text.")
    parser.add_argument("input", type=Path, help="Input .csv, .xlsx, .xls, .xlsm, .txt, or .tsv file.")
    parser.add_argument("--output-dir", type=Path, default=Path("unclaimed_funds_output"), help="Folder for HTML and CSV outputs.")
    parser.add_argument("-o", "--output", type=Path, default=None, help="Optional Excel workbook output path.")
    parser.add_argument("--xlsx", action="store_true", help="Also create an Excel workbook. If -o is omitted, uses the output folder.")
    parser.add_argument("--pasted-text", action="store_true", help="Treat the input file as pasted delimited text.")
    parser.add_argument("--fuzzy-threshold", type=int, default=90, help="Owner-name fuzzy match threshold from 0 to 100.")
    args = parser.parse_args()

    if not 0 <= args.fuzzy_threshold <= 100:
        raise SystemExit("--fuzzy-threshold must be between 0 and 100")

    raw_df = read_input(args.input, text_mode=args.pasted_text)
    lead_df = build_lead_dataframe(raw_df, fuzzy_threshold=args.fuzzy_threshold)
    report_path, sections = write_browser_outputs(lead_df, args.output_dir)
    print(f"Wrote {len(lead_df)} cleaned records to {args.output_dir}")
    print(f"HTML report: {report_path}")
    print("CSV files:")
    for section in sections:
        print(f"  {args.output_dir / section.filename}")

    if args.xlsx or args.output:
        xlsx_path = args.output or (args.output_dir / "unclaimed_funds_leads.xlsx")
        write_workbook(lead_df, xlsx_path)
        print(f"Excel workbook: {xlsx_path}")

    print("All generated leads are marked Needs Human Verification = Yes.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())



